"""
Extraction engine for IndustrialUseCases workbook.

Design principle: sheets are semi-structured narrative documents, not tables.
Row positions shift per sheet (variable node counts, variable decision counts),
so every section is located by searching for its ANCHOR LABEL text, never by
a fixed row number. Anchor labels were verified present in all 10 filled
sheets. Within a section, label->value pairs are read generically so that
case-specific fields (different KPI names, split root-cause fields, etc.)
are captured automatically without hardcoding a fixed field list.

Re-run this script any time the source Excel is updated:
    python3 extract.py <path_to_xlsx> <output_json_path>
"""
import sys
import json
import re
import openpyxl

CASE_SHEETS = [f"FIB-{i:03d}" for i in range(1, 11)]  # FIB-001..FIB-010 (filled sheets)


def cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def col_letter_to_idx(letter):
    return openpyxl.utils.column_index_from_string(letter)


B = col_letter_to_idx("B")
C = col_letter_to_idx("C")
G = col_letter_to_idx("G")
H = col_letter_to_idx("H")
I = col_letter_to_idx("I")


def find_row_with_label(ws, label, col=B, start=1, end=None):
    end = end or ws.max_row
    for r in range(start, end + 1):
        if cell(ws, r, col) == label:
            return r
    return None


def read_header_fields(ws):
    """Rows 2-6: Use case ID / Name / Industry / Brand / Scale (fixed, always present)."""
    return {
        "id": cell(ws, 2, C),
        "name": cell(ws, 3, C),
        "industry": cell(ws, 4, C),
        "brand_group": cell(ws, 5, C),
        "scale": cell(ws, 6, C),
    }


def read_node_descriptions(ws, start_row):
    """Col B/C node list under 'Baseline Operations (Normal Flow)' until blank."""
    nodes = []
    r = start_row
    while True:
        label = cell(ws, r, B)
        desc = cell(ws, r, C)
        if label is None:
            break
        nodes.append({"label": label, "description": desc})
        r += 1
    return nodes, r


def read_flow_table(ws, header_row):
    """Col G/H/I step table: Step / Process / Node. Runs until G is non-numeric."""
    steps = []
    r = header_row + 1
    while True:
        step = cell(ws, r, G)
        if not isinstance(step, (int, float)):
            break
        steps.append({
            "step": int(step),
            "process": cell(ws, r, H),
            "node": cell(ws, r, I),
        })
        r += 1
    return steps


def read_kv_section(ws, start_row, stop_labels, stop_contains=()):
    """Generic label(B) -> value(C) reader. Stops at blank-run, a stop label,
    or a label containing any of stop_contains (handles sheets with no blank-row
    gap before the next section)."""
    items = []
    r = start_row
    blank_streak = 0
    while r <= ws.max_row:
        label = cell(ws, r, B)
        if label in stop_labels:
            break
        if isinstance(label, str) and any(s in label for s in stop_contains):
            break
        if label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        items.append({"label": label, "value": cell(ws, r, C)})
        r += 1
    return items, r


def find_decision_blocks(ws, header_row):
    """Scan the header row for every cell containing 'Decision Analysis' -> (col, title)."""
    blocks = []
    for c in range(B, ws.max_column + 1):
        v = cell(ws, header_row, c)
        if isinstance(v, str) and "Decision Analysis" in v:
            blocks.append({"col": c, "title": v})
    return blocks


def read_decision_block(ws, label_col, decision_no_row):
    """Read one decision-analysis block: Decision No / Description / Action / rating rows / Best when.
    Table is transposed: label in `label_col`, one decision per subsequent column."""
    # Determine number of decision columns from the 'Decision No' row
    decision_cols = []
    c = label_col + 1
    while True:
        v = cell(ws, decision_no_row, c)
        if v is None:
            # allow a single gap then stop
            if cell(ws, decision_no_row, c + 1) is None:
                break
        else:
            decision_cols.append(c)
        c += 1
        if c > ws.max_column:
            break

    decisions = [{} for _ in decision_cols]
    r = decision_no_row
    blank_streak = 0
    while r <= ws.max_row:
        row_label = cell(ws, r, label_col)
        if row_label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        if isinstance(row_label, str) and ("Decision Analysis" in row_label or row_label in
                                            ("Decisions Impact Summary -", "Decisions Impact Summary - Self Analysis")):
            r += 1
            continue
        for idx, dc in enumerate(decision_cols):
            val = cell(ws, r, dc)
            if val is not None:
                decisions[idx][str(row_label)] = val
        r += 1
    # drop fully-empty decision entries
    decisions = [d for d in decisions if d]
    return decisions, r


def read_generic_table(ws, header_row, label_col=B):
    """Reads a table with a header row (Decision/Scenario + metric column names) and data rows below.
    Stops at blank-run of 2 or a new all-caps-ish section title."""
    headers = []
    c = label_col + 1
    while True:
        v = cell(ws, header_row, c)
        if v is None:
            break
        headers.append((c, str(v)))
        c += 1
    rows = []
    r = header_row + 1
    blank_streak = 0
    while r <= ws.max_row:
        row_label = cell(ws, r, label_col)
        if row_label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        entry = {"label": row_label}
        for c, hname in headers:
            v = cell(ws, r, c)
            if v is not None:
                entry[hname] = v
        rows.append(entry)
        r += 1
    return headers, rows, r


def compute_overall(score_rows):
    numeric_keys = ["Cost", "Resilience Gain", "Feasibility", "Speed", "Sustainability"]
    for row in score_rows:
        if "Overall" not in row or row.get("Overall") in (None, ""):
            vals = [row[k] for k in numeric_keys if isinstance(row.get(k), (int, float))]
            if vals:
                row["Overall"] = round(sum(vals) / len(vals), 2)
                row["_overall_calculated"] = True
    return score_rows


def read_sources(ws):
    sources = []
    A = col_letter_to_idx("A")
    header_row = None
    for r in range(1, ws.max_row + 1):
        if cell(ws, r, A) == "SOURCES":
            header_row = r
            break
    if header_row is None:
        return sources
    for r in range(header_row + 1, ws.max_row + 1):
        v = cell(ws, r, A)
        if v:
            sources.append(v)
    return sources


def extract_case(ws):
    header = read_header_fields(ws)
    if not header["id"]:
        return None  # placeholder / empty sheet

    row_baseline = find_row_with_label(ws, "Baseline Operations (Normal Flow)")
    nodes, _ = read_node_descriptions(ws, row_baseline + 1)
    flow_header_row = find_row_with_label(ws, "Step", col=G, start=row_baseline, end=row_baseline + 3)
    flow = read_flow_table(ws, flow_header_row) if flow_header_row else []

    row_kpi = find_row_with_label(ws, "Normal KPI Conditions")
    # row_kpi+1 is the "KPI"/"Value" header row itself
    kpis, after_kpi_row = read_kv_section(ws, row_kpi + 2, stop_labels={"Disruption State 1"})

    row_disruption_details = find_row_with_label(ws, "Disruption details", start=after_kpi_row)
    disruption, after_disruption_row = read_kv_section(
        ws, row_disruption_details + 1, stop_labels=set(), stop_contains=("Decision Analysis",)
    )

    # Decision Analysis blocks: find the header row (first row after disruption section
    # that contains a cell with "Decision Analysis" text, searched across all columns)
    r = after_disruption_row
    decision_header_row = None
    while r <= ws.max_row:
        row_vals = [cell(ws, r, c) for c in range(B, ws.max_column + 1)]
        if any(isinstance(v, str) and "Decision Analysis" in v for v in row_vals):
            decision_header_row = r
            break
        r += 1

    decision_blocks_out = []
    next_section_row = None
    if decision_header_row:
        blocks = find_decision_blocks(ws, decision_header_row)
        # "Decision No" row is usually 2 rows below the header (1 blank row between)
        for b in blocks:
            dn_row = find_row_with_label(ws, "Decision No", col=b["col"],
                                          start=decision_header_row, end=decision_header_row + 5)
            if dn_row is None:
                continue
            decisions, end_row = read_decision_block(ws, b["col"], dn_row)
            decision_blocks_out.append({
                "source": b["title"],
                "decisions": decisions,
            })
            next_section_row = max(next_section_row or 0, end_row)

    # Decisions Impact Summary (one or two side-by-side tables, similar treatment: search row for label)
    impact_row = find_row_with_label(ws, "Decisions Impact Summary -", start=next_section_row or 1)
    impact_summary = []
    row_after_impact = impact_row
    if impact_row:
        header_row = impact_row + 1
        _, rows, row_after_impact = read_generic_table(ws, header_row)
        impact_summary = rows
        # check for a second side-by-side impact table (Self Analysis) further right
        for c in range(C, ws.max_column + 1):
            v = cell(ws, header_row - 1, c)
            if isinstance(v, str) and "Decisions Impact Summary" in v:
                _, rows2, _ = read_generic_table(ws, header_row, label_col=c)
                impact_summary.append({"_second_table_start_col": c})
                impact_summary.extend(rows2)

    kpi_change_row = find_row_with_label(ws, "KPI Change Summary - Self Analyses", start=row_after_impact or 1)
    kpi_change_summary = []
    row_after_kpichange = kpi_change_row
    if kpi_change_row:
        header_row = kpi_change_row + 1
        _, rows, row_after_kpichange = read_generic_table(ws, header_row)
        kpi_change_summary = rows

    score_row = find_row_with_label(ws, "Decision Analysis Score - Self Analyses", start=row_after_kpichange or 1)
    decision_scores = []
    if score_row:
        # There can be TWO header rows (full question text, then short names) - use the row
        # immediately before the first data row (a row where the value col is numeric)
        header_row = score_row + 1
        r = header_row
        while r <= ws.max_row:
            v = cell(ws, r + 1, C)
            if isinstance(v, (int, float)):
                header_row = r
                break
            r += 1
        _, decision_scores, _ = read_generic_table(ws, header_row)
        decision_scores = compute_overall(decision_scores)

    sources = read_sources(ws)

    return {
        **header,
        "normal_flow_nodes": nodes,
        "normal_flow_steps": flow,
        "kpis": kpis,
        "disruption": disruption,
        "decision_blocks": decision_blocks_out,
        "decisions_impact_summary": impact_summary,
        "kpi_change_summary": kpi_change_summary,
        "decision_scores": decision_scores,
        "sources": sources,
    }


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/IndustrialUseCases-Vedang-Verified.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "cases.json"

    wb = openpyxl.load_workbook(src, data_only=True)
    cases = []
    errors = []
    for name in wb.sheetnames:
        if name == "Master Directory":
            continue
        ws = wb[name]
        try:
            case = extract_case(ws)
            if case:
                case["_sheet_name"] = name
                cases.append(case)
        except Exception as e:
            errors.append({"sheet": name, "error": str(e)})

    with open(out, "w") as f:
        json.dump({"cases": cases, "errors": errors}, f, indent=2, default=str)

    print(f"Extracted {len(cases)} cases -> {out}")
    if errors:
        print("ERRORS:", json.dumps(errors, indent=2))


if __name__ == "__main__":
    main()

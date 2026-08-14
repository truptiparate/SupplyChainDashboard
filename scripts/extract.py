"""
Extraction engine for the IndustrialUseCases workbook (Updated Template).

Design principle: sheets are semi-structured narrative documents, not tables.
Row positions shift per sheet (variable node counts, variable decision counts,
variable number of decision-analysis blocks), so every section is located by
searching for its ANCHOR LABEL text in column B, never by a fixed row number.

v2 changes (updated template):
  - New "Nature of usecase" field (Disruption | Improvement) under the header.
  - The old flat "Disruption details" section is now nested under a
    "Disruption State" wrapper label, and there's a NEW parallel "Improvement"
    section (Integration details / Problem / Proposed Solutions / etc.) that's
    only populated when Nature of usecase == "Improvement". Both are read;
    whichever branch has no data comes back with an empty items list.
  - The "Normal Flow" step table moved from columns G/H/I to columns B/C/D
    (directly under the node list), so it's now located dynamically by
    searching for a "Step" label rather than a fixed column.
  - Decision-analysis blocks are no longer identified by "Decision Analysis"
    text in a shared header row. Instead, every occurrence of the exact label
    "Decision No" (in any column) is treated as the anchor for a table, and
    the table is classified by its title (the nearest non-empty label above
    it in the same column):
        title contains "impact summary" -> row-oriented impact table
        title contains "score"          -> row-oriented score table
        anything else                   -> transposed decision block
    This handles both the old side-by-side layout and the new stacked
    (sequential) layout without hardcoding block names or positions.
  - SOURCES entries are listed in column B (not column A).

Re-run this script any time the source Excel is updated:
    python3 extract.py <path_to_xlsx> <output_json_path>
"""
import sys
import json
import openpyxl

SKIP_SHEET_SUBSTRINGS = ("directory",)  # case-insensitive; e.g. "Master Directory Template"


def cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def col_letter_to_idx(letter):
    return openpyxl.utils.column_index_from_string(letter)


A = col_letter_to_idx("A")
B = col_letter_to_idx("B")
C = col_letter_to_idx("C")


def find_row_with_label(ws, label, col=B, start=1, end=None):
    """Case-insensitive exact match search for `label` in `col`, scanning
    rows [start, end]. Case-insensitive because label casing is not
    consistent across sheets/authors (e.g. "Nature Of Usecase" vs
    "Nature of usecase")."""
    end = end or ws.max_row
    label_l = label.lower()
    for r in range(start, end + 1):
        v = cell(ws, r, col)
        if isinstance(v, str) and v.lower() == label_l:
            return r
    return None


def find_row_contains(ws, needle, col=B, start=1, end=None):
    """Case-insensitive substring search for `needle` in `col`."""
    end = end or ws.max_row
    needle = needle.lower()
    for r in range(start, end + 1):
        v = cell(ws, r, col)
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


def read_header_fields(ws):
    """Header block near the top: label in col B, value in col C.
    Located by label text (not fixed rows) so field order/spacing can shift."""
    def field(label):
        r = find_row_with_label(ws, label, start=1, end=15)
        return cell(ws, r, C) if r else None

    return {
        "id": field("Use case ID"),
        "name": field("Use case Name"),
        "industry": field("Industry Verticle"),
        "brand_group": field("Brand & Group"),
        "nature": field("Nature of usecase"),
        "scale": field("Scale"),
    }


def read_node_descriptions(ws, start_row):
    """Col B/C node list under 'Baseline Operations (Normal Flow)' until blank."""
    nodes = []
    r = start_row
    while True:
        label = cell(ws, r, B)
        desc = cell(ws, r, C)
        if label is None:
            break
        nodes.append({"label": label, "description": desc})
        r += 1
    return nodes, r


def read_flow_table(ws, header_row, step_col):
    """Step/Process/Node table. step_col is wherever 'Step' was found;
    Process and Node are assumed to be the next two columns to the right
    (matches both the old G/H/I layout and the new B/C/D layout)."""
    process_col = step_col + 1
    node_col = step_col + 2
    steps = []
    r = header_row + 1
    while True:
        step = cell(ws, r, step_col)
        if not isinstance(step, (int, float)):
            break
        steps.append({
            "step": int(step),
            "process": cell(ws, r, process_col),
            "node": cell(ws, r, node_col),
        })
        r += 1
    return steps


def read_kv_section(ws, start_row, stop_labels=(), stop_contains=()):
    """Generic label(B) -> value(C) reader. Stops at a blank-run of 2, an exact
    stop label, or a label containing any of stop_contains (handles sheets with
    no blank-row gap before the next section)."""
    items = []
    r = start_row
    blank_streak = 0
    while r <= ws.max_row:
        label = cell(ws, r, B)
        if label in stop_labels:
            break
        if isinstance(label, str) and any(s.lower() in label.lower() for s in stop_contains):
            break
        if label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        items.append({"label": label, "value": cell(ws, r, C)})
        r += 1
    return items, r


def read_generic_table(ws, header_row, label_col=B, max_row=None):
    """Reads a row-oriented table: header row has metric names starting at
    label_col+1; each data row below has a value in label_col (e.g. a decision
    number) plus one value per metric column. Stops at a blank-run of 2, or at
    max_row if given (used to avoid overrunning into the next anchored table
    when sections are separated by only a single blank row)."""
    headers = []
    c = label_col + 1
    while True:
        v = cell(ws, header_row, c)
        if v is None:
            break
        headers.append((c, str(v)))
        c += 1
    rows = []
    r = header_row + 1
    limit = min(max_row, ws.max_row) if max_row else ws.max_row
    blank_streak = 0
    while r <= limit:
        row_label = cell(ws, r, label_col)
        if row_label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        entry = {"label": row_label}
        for c, hname in headers:
            v = cell(ws, r, c)
            if v is not None:
                entry[hname] = v
        rows.append(entry)
        r += 1
    return headers, rows, r


def read_decision_block(ws, label_col, decision_no_row, max_row=None):
    """Transposed table: label in `label_col`, one decision per column to the
    right of it, starting at decision_no_row (the 'Decision No' row itself).
    `max_row` bounds reading so it can't overrun into the next anchored table
    when sections are separated by only a single blank row.

    Each returned decision dict includes "decision_no" (the number from the
    'Decision No' row itself, coerced to int where possible -- cells are
    inconsistently typed across the workbook, e.g. '1\\n' (str with a
    trailing newline) next to plain ints like 2, 3, 4)."""
    decision_cols = []
    c = label_col + 1
    while True:
        v = cell(ws, decision_no_row, c)
        if v is None:
            if cell(ws, decision_no_row, c + 1) is None:
                break
        else:
            decision_cols.append(c)
        c += 1
        if c > ws.max_column:
            break

    def decision_no_value(dc):
        v = cell(ws, decision_no_row, dc)
        if isinstance(v, str):
            v = v.strip()
            if v.isdigit():
                return int(v)
            return v
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v

    decisions = [{"decision_no": decision_no_value(dc)} for dc in decision_cols]
    r = decision_no_row
    limit = min(max_row, ws.max_row) if max_row else ws.max_row
    blank_streak = 0
    while r <= limit:
        row_label = cell(ws, r, label_col)
        if row_label is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        if isinstance(row_label, str) and "decision no" in row_label.lower():
            r += 1
            continue
        for idx, dc in enumerate(decision_cols):
            val = cell(ws, r, dc)
            if val is not None:
                decisions[idx][str(row_label)] = val
        r += 1
    # keep a decision only if it has real content beyond the decision number
    decisions = [d for d in decisions if len(d) > 1]
    return decisions, r


def nearest_title_above(ws, row, col, max_lookback=6):
    """Find the nearest non-empty label in `col` above `row` (used to name the
    table anchored at a 'Decision No' row — the title sits 1+ rows above it,
    with the exact gap varying by sheet)."""
    for r in range(row - 1, max(row - max_lookback, 0), -1):
        v = cell(ws, r, col)
        if v:
            return str(v), r
    return None, None


def compute_overall(score_rows):
    numeric_keys = ["Cost", "Resilience Gain", "Resilience gain", "Feasibility",
                     "Speed", "Speed of implementation", "Sustainability"]
    for row in score_rows:
        overall_key = next((k for k in row if k.lower().startswith("overall")), "Overall")
        if row.get(overall_key) in (None, ""):
            vals = [row[k] for k in numeric_keys if isinstance(row.get(k), (int, float))]
            if vals:
                row[overall_key] = round(sum(vals) / len(vals), 2)
                row["_overall_calculated"] = True
    return score_rows


def read_sources(ws):
    """SOURCES header + list, both in column B."""
    sources = []
    header_row = find_row_with_label(ws, "SOURCES", col=B)
    if header_row is None:
        return sources
    for r in range(header_row + 1, ws.max_row + 1):
        v = cell(ws, r, B)
        if v:
            sources.append(v)
    return sources


def find_all_decision_table_anchors(ws):
    """Every cell marking the start of a decision-related table, returned in
    reading order as dicts: {"header_row", "col", "title_lookup_row"}.

    Two distinct header texts mark these tables:
      - "Decision No"  -> transposed decision blocks
      - "Decision"     -> row-oriented impact-summary / score tables

    Both must be treated as anchors (not just "Decision No"), otherwise the
    impact-summary/score tables are invisible to the anchor scan: the bound
    for the PRECEDING transposed block then extends all the way to the next
    real "Decision No" block, so read_decision_block keeps reading straight
    through the impact-summary table's numbered rows (1, 2, 3...) as if they
    were more label/value pairs of the transposed block -- producing stray
    "1"/"2"/"3" keys on the wrong decisions. It also means the impact-summary
    and score tables themselves are never classified/read at all.

    Known template quirk: score-table headers are duplicated across two
    consecutive rows (slightly different metric-name casing) before the
    numeric data starts, e.g.:
        Decision | Cost | Resilience gain | ... | Overall (Average)
        Decision | Cost | Resilience Gain | ... | Overall
        1        | 4    | 4                | ...
    Only the SECOND (lower) row is the real header immediately above the
    data, so header_row points there. But the section title ("Decision
    Analysis Score - ...") sits above BOTH duplicate rows, so title_lookup_row
    is kept at the FIRST row -- nearest_title_above looks upward from
    title_lookup_row - 1, so pointing it at the first row skips past the
    duplicate header text and finds the real section title instead of just
    re-finding "Decision" again.
    """
    raw = []
    for r in range(1, ws.max_row + 1):
        for c in range(B, ws.max_column + 1):
            v = cell(ws, r, c)
            if isinstance(v, str) and v.strip().lower() in ("decision no", "decision"):
                raw.append((r, c, v.strip().lower()))
    raw.sort()

    anchors = []
    i = 0
    while i < len(raw):
        row, col, kind = raw[i]
        title_lookup_row = row
        # Collapse a run of duplicate "Decision" header rows (same column,
        # immediately consecutive rows) into one anchor at the LAST row of
        # the run, while keeping the FIRST row for title lookup.
        while (
            i + 1 < len(raw)
            and raw[i + 1][1] == col
            and raw[i + 1][0] == row + 1
            and kind == "decision"
            and raw[i + 1][2] == "decision"
        ):
            i += 1
            row = raw[i][0]
        anchors.append({"header_row": row, "col": col, "title_lookup_row": title_lookup_row})
        i += 1
    return anchors


def extract_case(ws):
    header = read_header_fields(ws)
    if not header["id"]:
        return None  # placeholder / empty sheet

    row_baseline = find_row_with_label(ws, "Baseline Operations (Normal Flow)")
    if row_baseline is None:
        return {**header, "_error": "missing 'Baseline Operations (Normal Flow)' section"}
    nodes, after_nodes_row = read_node_descriptions(ws, row_baseline + 1)

    flow_header_row = find_row_with_label(ws, "Step", col=B, start=after_nodes_row, end=after_nodes_row + 6)
    flow = read_flow_table(ws, flow_header_row, step_col=B) if flow_header_row else []
    after_flow_row = flow_header_row + len(flow) + 1 if flow_header_row else after_nodes_row

    row_kpi = find_row_with_label(ws, "Normal KPI Conditions", start=after_flow_row)
    kpis, after_kpi_row = ([], after_flow_row)
    if row_kpi:
        # row_kpi+1 is the "KPI"/"Value" header row itself.
        # stop_contains must cover whatever section can immediately follow
        # the KPI table on EITHER nature branch: Disruption sheets are
        # followed by "Disruption State ...", Improvement sheets are
        # followed by "Technology" (optional) then "Integration steps" /
        # "Integration details" -- none of which contain the word
        # "Improvement", so "Integration"/"Technology" must be listed
        # explicitly or the KPI read overruns straight through the rest
        # of the sheet on Improvement-nature cases.
        kpis, after_kpi_row = read_kv_section(
            ws, row_kpi + 2,
            stop_contains=("Disruption", "Improvement", "Integration", "Technology"),
        )

    # --- Disruption State branch ---
    disruption = []
    after_disruption_row = after_kpi_row
    row_disruption_details = find_row_with_label(ws, "Disruption details", start=after_kpi_row)
    if row_disruption_details:
        disruption, after_disruption_row = read_kv_section(
            ws, row_disruption_details + 1, stop_contains=("Improvement", "Decision")
        )

    # --- Improvement branch (parallel section; populated only for Improvement-nature cases) ---
    # NOTE: earlier versions gated this on finding a literal "Improvement"
    # wrapper label before "Integration details" -- that wrapper label does
    # not exist in this template (verified against all Improvement-nature
    # sheets), so that gate always failed and this branch was always empty.
    # "Integration details" is searched for directly instead.
    improvement = []
    after_improvement_row = after_disruption_row
    row_integration_details = find_row_with_label(
        ws, "Integration details", start=after_disruption_row
    )
    if row_integration_details:
        improvement, after_improvement_row = read_kv_section(
            ws, row_integration_details + 1, stop_contains=("Decision",)
        )
    # drop empty rows (fields with no value are just unused template rows for this case's nature)
    disruption = [d for d in disruption if d["value"] is not None]
    improvement = [d for d in improvement if d["value"] is not None]

    # --- Decision-related tables: every 'Decision No' anchor, classified by its title.
    # Each table's read is bounded by the row of the NEXT anchor (or SOURCES, if that
    # comes first) so it can't overrun into a neighboring table when sections are
    # separated by only a single blank row.
    sources_row = find_row_with_label(ws, "SOURCES", col=B)
    anchors = [a for a in find_all_decision_table_anchors(ws) if a["header_row"] >= after_improvement_row]

    decision_blocks_out = []
    impact_summary = []
    decision_scores = []
    for i, anchor in enumerate(anchors):
        row, col = anchor["header_row"], anchor["col"]
        next_anchor_row = anchors[i + 1]["header_row"] if i + 1 < len(anchors) else None
        candidates = [r for r in (next_anchor_row, sources_row, ws.max_row + 1) if r]
        bound = min(candidates) - 1

        title, _ = nearest_title_above(ws, anchor["title_lookup_row"], col)
        title_l = (title or "").lower()
        if "impact summary" in title_l:
            _, rows, _ = read_generic_table(ws, row, label_col=col, max_row=bound)
            rows = [r for r in rows if len(r) > 1]  # drop stray section-title rows with no metric data
            impact_summary.extend(rows)
        elif "score" in title_l:
            _, rows, _ = read_generic_table(ws, row, label_col=col, max_row=bound)
            rows = [r for r in rows if len(r) > 1]
            decision_scores.extend(compute_overall(rows))
        else:
            decisions, _ = read_decision_block(ws, col, row, max_row=bound)
            decision_blocks_out.append({"source": title or "Decisions", "decisions": decisions})

    kpi_change_row = find_row_contains(ws, "KPI Change Summary")
    kpi_change_summary = []
    if kpi_change_row:
        _, kpi_change_summary, _ = read_generic_table(ws, kpi_change_row + 1)

    sources = read_sources(ws)

    return {
        **header,
        "normal_flow_nodes": nodes,
        "normal_flow_steps": flow,
        "kpis": kpis,
        "disruption": disruption,
        "improvement": improvement,
        "decision_blocks": decision_blocks_out,
        "decisions_impact_summary": impact_summary,
        "kpi_change_summary": kpi_change_summary,
        "decision_scores": decision_scores,
        "sources": sources,
    }


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/UpdatedTemplateNew.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "cases.json"

    wb = openpyxl.load_workbook(src, data_only=True)
    cases = []
    errors = []
    for name in wb.sheetnames:
        if any(s in name.lower() for s in SKIP_SHEET_SUBSTRINGS):
            continue
        ws = wb[name]
        try:
            case = extract_case(ws)
            if case:
                case["_sheet_name"] = name
                cases.append(case)
        except Exception as e:
            errors.append({"sheet": name, "error": str(e)})

    with open(out, "w") as f:
        json.dump({"cases": cases, "errors": errors}, f, indent=2, default=str)

    print(f"Extracted {len(cases)} cases -> {out}")
    if errors:
        print("ERRORS:", json.dumps(errors, indent=2))


if __name__ == "__main__":
    main()
